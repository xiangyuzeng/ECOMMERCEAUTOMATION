"""
Excel workbook generator — 8-tab 运营方案.xlsx using openpyxl.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import logging
import math

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
BODY_FONT = Font(name='Arial', size=10, color='2C3E50')
ALT_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
INPUT_FONT = Font(name='Arial', size=10, color='0000FF')
INPUT_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Conditional formatting fills
RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_body(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT
            if (r - start_row) % 2 == 1:
                cell.fill = ALT_FILL


def _auto_width(ws, max_col, max_row):
    for col in range(1, max_col + 1):
        max_len = 10
        for row in range(1, min(max_row + 1, 100)):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        width = min(max(max_len + 2, 10), 45)
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_input_cell(ws, row, col, value=None):
    """Style a user-editable cell with blue font + yellow fill."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = THIN_BORDER
    return cell


def _safe_val(val, default='N/A'):
    """Return value or default if None/NaN."""
    if val is None:
        return default
    try:
        if math.isnan(val):
            return default
    except (TypeError, ValueError):
        pass
    return val


def write_workbook(filepath, competitor_matrix, keyword_library,
                   ads_data, pricing_data, traffic_sources,
                   gap_analysis, file_log, config):
    """Generate 8-tab Excel workbook."""
    wb = openpyxl.Workbook()

    # Tab 1: 竞品分析
    _write_competitors_tab(wb, competitor_matrix, config)

    # Tab 2: 产品清单
    _write_products_tab(wb, pricing_data, config)

    # Tab 3: 词库整理
    _write_keywords_tab(wb, keyword_library)

    # Tab 4: 广告指标监测
    _write_ads_tab(wb, ads_data)

    # Tab 5: 定价策略
    _write_pricing_tab(wb, pricing_data)

    # Tab 6: 流量入口
    _write_traffic_tab(wb, traffic_sources)

    # Tab 7: 关键词Gap分析
    _write_gap_tab(wb, gap_analysis, config)

    # Tab 8: 数据源日志
    _write_log_tab(wb, file_log)

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(filepath)
    logger.info(f"Workbook saved: {filepath}")


def _write_competitors_tab(wb, matrix, config):
    """Tab 1: 竞品分析 — vertical comparison matrix."""
    ws = wb.create_sheet('竞品分析')

    metrics = [
        ('图片', 'image_url'),
        ('品牌', 'brand'),
        ('ASIN', 'asin'),
        ('价格', 'price'),
        ('排名', '_bsr_combined'),
        ('标题', 'title'),
        ('前五核心流量词', 'top_keywords'),
        ('评论数', 'ratings_count'),
        ('月销量', 'monthly_sales'),
        ('月销售额', 'monthly_revenue'),
        ('上架日期', 'launch_date'),
        ('变体数', 'variation_count'),
        ('FBA毛利率', 'fba_margin'),
        ('流量关键词数', 'keyword_count'),
    ]

    # Pre-compute combined BSR string for each product
    if matrix:
        for prod in matrix:
            cat_bsr = prod.get('category_bsr')
            sub_bsr = prod.get('subcategory_bsr')
            parts = []
            if cat_bsr and cat_bsr != 'N/A' and not (isinstance(cat_bsr, float) and math.isnan(cat_bsr)):
                parts.append(f"类目#{int(cat_bsr) if isinstance(cat_bsr, (int, float)) else cat_bsr}")
            if sub_bsr and sub_bsr != 'N/A' and not (isinstance(sub_bsr, float) and math.isnan(sub_bsr)):
                parts.append(f"子类#{int(sub_bsr) if isinstance(sub_bsr, (int, float)) else sub_bsr}")
            prod['_bsr_combined'] = ' / '.join(parts) if parts else 'N/A'

    # Header row: metric | my product | competitors
    ws.cell(row=1, column=1, value='指标')
    if matrix:
        for i, prod in enumerate(matrix):
            label = prod.get('brand', '') + (' (我)' if prod.get('is_mine') else '')
            ws.cell(row=1, column=i + 2, value=label)

    # Data rows
    for r, (label, key) in enumerate(metrics, start=2):
        ws.cell(row=r, column=1, value=label)
        if matrix:
            for i, prod in enumerate(matrix):
                val = prod.get(key, 'N/A')
                ws.cell(row=r, column=i + 2, value=_safe_val(val))

    max_col = len(matrix) + 1 if matrix else 2
    max_row = len(metrics) + 1
    _style_header(ws, 1, max_col)
    _style_body(ws, 2, max_row, max_col)
    _auto_width(ws, max_col, max_row)
    ws.freeze_panes = 'B2'


def _write_products_tab(wb, pricing_data, config):
    """Tab 2: 产品清单 — cost model + variant sales."""
    ws = wb.create_sheet('产品清单')

    scenarios = pricing_data.get('scenarios', [])
    cost = pricing_data.get('cost_inputs', {})

    # Section A: Cost model
    ws.cell(row=1, column=1, value='成本项')
    for i, sc in enumerate(scenarios):
        ws.cell(row=1, column=i + 2, value=sc['scenario'])

    cost_rows = [
        ('售价', 'price'),
        ('成品成本', 'unit_cost'),
        ('包装', 'packaging'),
        ('头程运费', 'inbound_shipping'),
        ('Amazon佣金(17%)', None),  # Formula
        ('FBA配送费', 'fba_fee'),
        ('月仓储费', 'storage_fee'),
        ('广告费(CPA)', None),  # Formula
        ('退货损失', None),  # Formula
        ('总成本', None),
        ('毛利润', None),
        ('毛利率', None),
    ]

    for r, (label, key) in enumerate(cost_rows, start=2):
        ws.cell(row=r, column=1, value=label)
        for i, sc in enumerate(scenarios):
            col = i + 2
            col_letter = get_column_letter(col)

            if label == '售价':
                ws.cell(row=r, column=col, value=sc['price'])
            elif label == '成品成本':
                cell = _write_input_cell(ws, r, col, sc['unit_cost'])
            elif label == '包装':
                cell = _write_input_cell(ws, r, col, sc['packaging'])
            elif label == '头程运费':
                cell = _write_input_cell(ws, r, col, sc['inbound_shipping'])
            elif label == 'Amazon佣金(17%)':
                # Formula: price * 0.17
                ws.cell(row=r, column=col, value=f'={col_letter}2*0.17')
            elif label == 'FBA配送费':
                ws.cell(row=r, column=col, value=sc['fba_fee'])
            elif label == '月仓储费':
                ws.cell(row=r, column=col, value=sc['storage_fee'])
            elif label == '广告费(CPA)':
                # Formula: price * ad_rate
                ws.cell(row=r, column=col, value=f'={col_letter}2*{sc["ad_rate"]}')
            elif label == '退货损失':
                # Formula: price * return_rate
                ws.cell(row=r, column=col, value=f'={col_letter}2*{sc["return_rate"]}')
            elif label == '总成本':
                # Sum of cost rows (rows 3-10)
                ws.cell(row=r, column=col, value=f'=SUM({col_letter}3:{col_letter}10)')
            elif label == '毛利润':
                # Price - total cost
                ws.cell(row=r, column=col, value=f'={col_letter}2-{col_letter}11')
            elif label == '毛利率':
                # Margin / price
                ws.cell(row=r, column=col, value=f'={col_letter}12/{col_letter}2')
                ws.cell(row=r, column=col).number_format = '0.0%'

    max_col = len(scenarios) + 1
    _style_header(ws, 1, max_col)
    _style_body(ws, 2, 13, max_col)

    # Section B: Variant sales
    variants = pricing_data.get('variants', [])
    if variants:
        start_row = 16
        ws.cell(row=start_row - 1, column=1, value='变体销售明细')
        ws.cell(row=start_row - 1, column=1).font = Font(name='Arial', size=12, bold=True, color='1F4E79')

        headers = ['ASIN', 'SKU', '销量', '销售额', 'Sessions', 'CVR', '平均售价']
        for c, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=c, value=h)
        _style_header(ws, start_row, len(headers))

        for r, v in enumerate(variants, start=start_row + 1):
            ws.cell(row=r, column=1, value=v['asin'])
            ws.cell(row=r, column=2, value=v.get('sku', ''))
            ws.cell(row=r, column=3, value=v['units_ordered'])
            ws.cell(row=r, column=4, value=v['revenue'])
            ws.cell(row=r, column=4).number_format = '$#,##0.00'
            ws.cell(row=r, column=5, value=v['sessions'])
            ws.cell(row=r, column=6, value=v['cvr'])
            ws.cell(row=r, column=6).number_format = '0.0%'
            ws.cell(row=r, column=7, value=v['avg_price'])
            ws.cell(row=r, column=7).number_format = '$#,##0.00'

        _style_body(ws, start_row + 1, start_row + len(variants), len(headers))

    _auto_width(ws, max(max_col, 7), start_row + len(variants) if variants else 13)
    ws.freeze_panes = 'B2'


def _write_keywords_tab(wb, keyword_library):
    """Tab 3: 词库整理 — keyword library with classification."""
    ws = wb.create_sheet('词库整理')

    headers = ['关键词', '一级分类', '二级分类', '搜索层级', '用途',
               '月搜索量', '搜索频率排名', '购买率', 'CPA', 'CPC', '自然排名',
               '广告排名', '广告花费', '广告ACoS', '数据来源']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    if keyword_library is not None and not keyword_library.empty:
        for r, (_, row) in enumerate(keyword_library.iterrows(), start=2):
            ws.cell(row=r, column=1, value=_safe_val(row.get('keyword')))
            ws.cell(row=r, column=2, value=_safe_val(row.get('一级分类')))
            ws.cell(row=r, column=3, value=_safe_val(row.get('二级分类')))
            ws.cell(row=r, column=4, value=_safe_val(row.get('搜索层级')))
            ws.cell(row=r, column=5, value=_safe_val(row.get('用途')))
            ws.cell(row=r, column=6, value=_safe_val(row.get('monthly_searches'), 0))
            ws.cell(row=r, column=7, value=_safe_val(row.get('aba_rank')))
            ws.cell(row=r, column=8, value=_safe_val(row.get('purchase_rate')))
            if row.get('purchase_rate') and row.get('purchase_rate') != 'N/A':
                ws.cell(row=r, column=8).number_format = '0.00%'
            ws.cell(row=r, column=9, value=_safe_val(row.get('cpa')))
            if row.get('cpa') and row.get('cpa') != 'N/A':
                ws.cell(row=r, column=9).number_format = '$#,##0.00'
            ws.cell(row=r, column=10, value=_safe_val(row.get('ppc_bid')))
            if row.get('ppc_bid') and row.get('ppc_bid') != 'N/A':
                ws.cell(row=r, column=10).number_format = '$#,##0.00'
            ws.cell(row=r, column=11, value=_safe_val(row.get('organic_rank')))
            ws.cell(row=r, column=12, value=_safe_val(row.get('sponsored_rank')))
            ws.cell(row=r, column=13, value=_safe_val(row.get('ad_spend')))
            ws.cell(row=r, column=14, value=_safe_val(row.get('ad_acos')))
            ws.cell(row=r, column=15, value=_safe_val(row.get('data_source')))

        max_row = len(keyword_library) + 1
        _style_body(ws, 2, max_row, len(headers))
    else:
        max_row = 1

    _auto_width(ws, len(headers), min(max_row, 50))
    ws.freeze_panes = 'B2'


def _write_ads_tab(wb, ads_data):
    """Tab 4: 广告指标监测 — weekly heatmap + search term + campaign tables."""
    ws = wb.create_sheet('广告指标监测')

    heatmap = ads_data.get('heatmap', {})
    keywords = heatmap.get('keywords', [])
    weeks = heatmap.get('weeks', [])
    grid = heatmap.get('grid', {})
    sub_metrics = ['曝光', '点击', '订单', '位置']

    # Section 1: Weekly heatmap
    # Header row: 关键词 | 指标 | week1 | week2 | ...
    ws.cell(row=1, column=1, value='关键词')
    ws.cell(row=1, column=2, value='指标')
    for wi, week in enumerate(weeks):
        ws.cell(row=1, column=wi + 3, value=week)
    heatmap_max_col = max(len(weeks) + 2, 3)
    _style_header(ws, 1, heatmap_max_col)

    current_row = 2
    for kw in keywords:
        kw_grid = grid.get(kw, {})
        for si, metric in enumerate(sub_metrics):
            r = current_row + si
            if si == 0:
                ws.cell(row=r, column=1, value=kw)
                # Merge keyword cells across 4 sub-rows
                if len(sub_metrics) > 1:
                    ws.merge_cells(start_row=r, start_column=1, end_row=r + len(sub_metrics) - 1, end_column=1)
            ws.cell(row=r, column=2, value=metric)
            metric_data = kw_grid.get(metric, {})
            for wi, week in enumerate(weeks):
                val = metric_data.get(week)
                ws.cell(row=r, column=wi + 3, value=_safe_val(val, 'N/A'))
        current_row += len(sub_metrics)

    heatmap_end_row = current_row - 1
    if heatmap_end_row >= 2:
        _style_body(ws, 2, heatmap_end_row, heatmap_max_col)

    # Section 2: Search term performance
    st_data = ads_data.get('search_term_summary', [])
    if st_data:
        st_start = heatmap_end_row + 3
        ws.cell(row=st_start - 1, column=1, value='搜索词表现')
        ws.cell(row=st_start - 1, column=1).font = Font(name='Arial', size=12, bold=True, color='1F4E79')

        st_headers = ['关键词', '曝光', '点击', 'CTR', 'CPC', '花费',
                       '销售额', 'ACoS', '订单', '转化率']
        for c, h in enumerate(st_headers, 1):
            ws.cell(row=st_start, column=c, value=h)
        _style_header(ws, st_start, len(st_headers))

        for r, row in enumerate(st_data, start=st_start + 1):
            ws.cell(row=r, column=1, value=row.get('keyword', ''))
            ws.cell(row=r, column=2, value=row.get('impressions', 0))
            ws.cell(row=r, column=3, value=row.get('clicks', 0))
            ws.cell(row=r, column=4, value=row.get('ctr', 0))
            ws.cell(row=r, column=4).number_format = '0.00%'
            ws.cell(row=r, column=5, value=row.get('cpc', 0))
            ws.cell(row=r, column=5).number_format = '$#,##0.00'
            ws.cell(row=r, column=6, value=row.get('spend', 0))
            ws.cell(row=r, column=6).number_format = '$#,##0.00'
            ws.cell(row=r, column=7, value=row.get('sales', 0))
            ws.cell(row=r, column=7).number_format = '$#,##0.00'
            ws.cell(row=r, column=8, value=row.get('acos', 0))
            ws.cell(row=r, column=8).number_format = '0.0%'
            ws.cell(row=r, column=9, value=row.get('orders', 0))
            ws.cell(row=r, column=10, value=row.get('cvr', 0))
            ws.cell(row=r, column=10).number_format = '0.0%'

        st_end = st_start + len(st_data)
        _style_body(ws, st_start + 1, st_end, len(st_headers))
    else:
        st_end = heatmap_end_row + 1

    # Section 3: Campaign summary
    campaigns = ads_data.get('campaign_summary', [])
    if campaigns:
        camp_start = st_end + 3
        ws.cell(row=camp_start - 1, column=1, value='广告活动概览')
        ws.cell(row=camp_start - 1, column=1).font = Font(name='Arial', size=12, bold=True, color='1F4E79')

        camp_headers = ['广告活动', '状态', '日预算', '曝光', '点击', '花费', '销售额', 'ACoS', '订单']
        for c, h in enumerate(camp_headers, 1):
            ws.cell(row=camp_start, column=c, value=h)
        _style_header(ws, camp_start, len(camp_headers))

        for r, camp in enumerate(campaigns, start=camp_start + 1):
            ws.cell(row=r, column=1, value=camp.get('campaign', ''))
            ws.cell(row=r, column=2, value=camp.get('status', ''))
            ws.cell(row=r, column=3, value=camp.get('budget', 0))
            ws.cell(row=r, column=3).number_format = '$#,##0.00'
            ws.cell(row=r, column=4, value=camp.get('impressions', 0))
            ws.cell(row=r, column=5, value=camp.get('clicks', 0))
            ws.cell(row=r, column=6, value=camp.get('spend', 0))
            ws.cell(row=r, column=6).number_format = '$#,##0.00'
            ws.cell(row=r, column=7, value=camp.get('sales', 0))
            ws.cell(row=r, column=7).number_format = '$#,##0.00'
            ws.cell(row=r, column=8, value=camp.get('acos', 0))
            ws.cell(row=r, column=8).number_format = '0.0%'
            ws.cell(row=r, column=9, value=camp.get('orders', 0))

        _style_body(ws, camp_start + 1, camp_start + len(campaigns), len(camp_headers))

    _auto_width(ws, heatmap_max_col, min(heatmap_end_row, 50))
    ws.freeze_panes = 'C2'


def _write_pricing_tab(wb, pricing_data):
    """Tab 5: 定价策略 — per-variant P&L table."""
    ws = wb.create_sheet('定价策略')

    headers = ['ASIN', 'SKU', '销量', '销售额', '平均售价',
               '单位成本', '采购占比', '单位头程', '头程占比',
               '单位配送费', '配送费占比', '类目佣金', '佣金占比',
               '月仓储费', '月仓储费占比', '广告花费', 'ACoS', 'TACoS']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    variants = pricing_data.get('variants', [])
    for r, v in enumerate(variants, start=2):
        c = get_column_letter
        ws.cell(row=r, column=1, value=v['asin'])
        ws.cell(row=r, column=2, value=v.get('sku', ''))
        ws.cell(row=r, column=3, value=v['units_ordered'])
        ws.cell(row=r, column=4, value=v['revenue'])
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

        # 平均售价
        ws.cell(row=r, column=5, value=v['avg_price'] if v.get('avg_price') is not None else 0)
        ws.cell(row=r, column=5).number_format = '$#,##0.00'

        # 单位成本 (editable)
        _write_input_cell(ws, r, 6, v['unit_cost'])
        ws.cell(row=r, column=6).number_format = '$#,##0.00'

        # 采购占比 = unit_cost / avg_price (Excel formula)
        ws.cell(row=r, column=7, value=f'=F{r}/E{r}')
        ws.cell(row=r, column=7).number_format = '0.0%'

        # 单位头程 (editable)
        _write_input_cell(ws, r, 8, v['inbound_shipping'])
        ws.cell(row=r, column=8).number_format = '$#,##0.00'

        # 头程占比
        ws.cell(row=r, column=9, value=f'=H{r}/E{r}')
        ws.cell(row=r, column=9).number_format = '0.0%'

        # 单位配送费
        ws.cell(row=r, column=10, value=v['fba_fee'])
        ws.cell(row=r, column=10).number_format = '$#,##0.00'

        # 配送费占比
        ws.cell(row=r, column=11, value=f'=J{r}/E{r}')
        ws.cell(row=r, column=11).number_format = '0.0%'

        # 类目佣金
        ws.cell(row=r, column=12, value=v['referral_fee'])
        ws.cell(row=r, column=12).number_format = '$#,##0.00'

        # 佣金占比
        ws.cell(row=r, column=13, value=f'=L{r}/E{r}')
        ws.cell(row=r, column=13).number_format = '0.0%'

        # 月仓储费
        ws.cell(row=r, column=14, value=v['storage_fee'])
        ws.cell(row=r, column=14).number_format = '$#,##0.00'

        # 月仓储费占比
        ws.cell(row=r, column=15, value=f'=N{r}/E{r}')
        ws.cell(row=r, column=15).number_format = '0.0%'

        # 广告花费
        ws.cell(row=r, column=16, value=v['ad_spend'])
        ws.cell(row=r, column=16).number_format = '$#,##0.00'

        # ACoS
        ws.cell(row=r, column=17, value=v['acos'])
        ws.cell(row=r, column=17).number_format = '0.0%'

        # TACoS
        ws.cell(row=r, column=18, value=v['tacos'])
        ws.cell(row=r, column=18).number_format = '0.0%'

        # Mark estimated rows with italic gray SKU
        if v.get('is_estimated'):
            ws.cell(row=r, column=2).font = Font(name='Arial', size=9, italic=True, color='999999')

    max_row = len(variants) + 1
    if max_row > 1:
        _style_body(ws, 2, max_row, len(headers))

    _auto_width(ws, len(headers), max_row)
    ws.freeze_panes = 'C2'


def _write_traffic_tab(wb, traffic_sources):
    """Tab 6: 流量入口 — traffic channel strategy."""
    ws = wb.create_sheet('流量入口')

    headers = ['流量入口', '流量来源', '方案']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for r, src in enumerate(traffic_sources or [], start=2):
        ws.cell(row=r, column=1, value=src.get('channel', ''))
        ws.cell(row=r, column=2, value=src.get('source', ''))
        ws.cell(row=r, column=3, value=src.get('strategy', ''))

    max_row = len(traffic_sources or []) + 1
    if max_row > 1:
        _style_body(ws, 2, max_row, len(headers))

    _auto_width(ws, len(headers), max_row)
    ws.freeze_panes = 'A2'


def _write_gap_tab(wb, gap_analysis, config=None):
    """Tab 7: 关键词Gap分析."""
    ws = wb.create_sheet('关键词Gap分析')

    # Dynamic competitor brand label from config
    comp_brand = 'Competitor'
    if config:
        competitors = config.get('competitors', {})
        for slot in ['C2', 'C1', 'C3', 'C4']:
            brand = competitors.get(slot, {}).get('brand', '') if isinstance(competitors.get(slot), dict) else ''
            if brand:
                comp_brand = brand
                break

    comp_rank_col = f'{comp_brand}_rank'

    headers = ['关键词', '我方排名', f'{comp_brand}排名', 'Gap类型',
               '月搜索量', '购买率', 'PPC竞价', '优先级分数', '建议操作']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    if gap_analysis is not None and not gap_analysis.empty:
        for r, (_, row) in enumerate(gap_analysis.iterrows(), start=2):
            ws.cell(row=r, column=1, value=_safe_val(row.get('keyword')))
            ws.cell(row=r, column=2, value=_safe_val(row.get('my_rank')))
            ws.cell(row=r, column=3, value=_safe_val(row.get(comp_rank_col)))
            ws.cell(row=r, column=4, value=_safe_val(row.get('gap_type')))
            ws.cell(row=r, column=5, value=_safe_val(row.get('monthly_searches'), 0))
            ws.cell(row=r, column=6, value=_safe_val(row.get('purchase_rate')))
            if row.get('purchase_rate') and row.get('purchase_rate') != 'N/A':
                ws.cell(row=r, column=6).number_format = '0.00%'
            ws.cell(row=r, column=7, value=_safe_val(row.get('ppc_bid')))
            if row.get('ppc_bid') and row.get('ppc_bid') != 'N/A':
                ws.cell(row=r, column=7).number_format = '$#,##0.00'
            ws.cell(row=r, column=8, value=_safe_val(row.get('priority_score'), 0))
            ws.cell(row=r, column=8).number_format = '#,##0.0'
            ws.cell(row=r, column=9, value=_safe_val(row.get('recommended_action')))

        max_row = len(gap_analysis) + 1
        _style_body(ws, 2, max_row, len(headers))

        # Conditional formatting on gap_type column (D)
        gap_col = f'D2:D{max_row}'
        ws.conditional_formatting.add(gap_col, CellIsRule(
            operator='equal', formula=['"MISSING"'], fill=RED_FILL))
        ws.conditional_formatting.add(gap_col, CellIsRule(
            operator='equal', formula=['"CATCHUP"'], fill=ORANGE_FILL))
        ws.conditional_formatting.add(gap_col, CellIsRule(
            operator='equal', formula=['"DEFEND"'], fill=GREEN_FILL))
    else:
        max_row = 1

    _auto_width(ws, len(headers), min(max_row, 50))
    ws.freeze_panes = 'B2'


def _write_log_tab(wb, file_log):
    """Tab 8: 数据源日志 — file audit trail."""
    ws = wb.create_sheet('数据源日志')

    headers = ['时间戳', '源文件', '源类型', '模块', '记录数',
               '对应Tab', '数据范围', '数据质量', '备注']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header(ws, 1, len(headers))

    for r, entry in enumerate(file_log or [], start=2):
        ws.cell(row=r, column=1, value=entry.get('timestamp', ''))
        ws.cell(row=r, column=2, value=entry.get('source_file', ''))
        ws.cell(row=r, column=3, value=entry.get('source_type', ''))
        ws.cell(row=r, column=4, value=entry.get('module', ''))
        ws.cell(row=r, column=5, value=entry.get('records', 0))
        ws.cell(row=r, column=6, value=entry.get('feeds_tab', ''))
        ws.cell(row=r, column=7, value=entry.get('date_range', ''))
        ws.cell(row=r, column=8, value=entry.get('data_quality', ''))
        ws.cell(row=r, column=9, value=entry.get('notes', ''))

    max_row = len(file_log or []) + 1
    if max_row > 1:
        _style_body(ws, 2, max_row, len(headers))

    _auto_width(ws, len(headers), max_row)
    ws.freeze_panes = 'A2'
